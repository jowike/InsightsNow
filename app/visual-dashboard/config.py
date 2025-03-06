import os
import yaml
from dateutil import relativedelta
import pandas as pd
from openpyxl import load_workbook
from pages.icons.hero import ICON

def __load_yaml(file_path):
    with open(file_path, "r") as file:
        return file.read()
    
def format_diff(diff_value):
    if diff_value > 0:
        return "text-success fw-bolder", ICON.UP_ARROW.XS
    elif diff_value < 0:
        return "text-danger fw-bolder", ICON.DOWN_ARROW.XS
    return "text fw-bolder", ICON.CHEVRON_UP_DOWN


def load_predictions(
        file_path:str="../analytical-backend/data/08_reporting/dash_data_model.xlsx",
        # type: str,
        ):

    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        assert "Nowcast Browser – Base" in wb.sheetnames and "Nowcast Browser – Header" in wb.sheetnames
        df = pd.read_excel(file_path, sheet_name="Nowcast Browser – Base")

        topBar = pd.read_excel(file_path, sheet_name="Nowcast Browser – Header")

        header = {}
        for k, v in zip(topBar["Banner"], topBar["Value"]):
            header[k] = v

        # if type == "base":
        #     assert "Nowcast Browser – Base" in wb.sheetnames
        #     df = pd.read_excel(file_path, sheet_name="Nowcast Browser – Base")
        # elif type == "adj":
        #     assert "Nowcast Browser – Adjusted" in wb.sheetnames
        #     df = pd.read_excel(file_path, sheet_name="Nowcast Browser – Adjusted")
        # else: raise ValueError("Invalid type. Please choose 'base' or 'adj'.")

        df["Reference Date"] = df["Reference Date"].astype(str).replace(r"-\d{2}$", "", regex=True)

        return {
            "labels": [label if index % 3 == 0 else "" for index, label in enumerate(df["Reference Date"])],
            "series": [df[c].tolist() for c in df.columns if c != "Reference Date"]
        }, header
    return None, header


def load_cards(
        file_path:str="../analytical-backend/data/08_reporting/dash_data_model.xlsx",
        ):
    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        assert "Cards" in wb.sheetnames

        df = pd.read_excel(file_path, sheet_name="Cards")

        d = {}
        for r in df.iterrows():
            d[r[1]["Card"]] = {
                "Value": r[1]["Value"],
                "Since Last Month": r[1]["Since Last Month"],
                # "Reference Period": r[1]["Reference Period"],
                # "Region": r[1]["Region"],
                "Prediction Range": r[1]["Prediction Range"],
            }
        return d
    return None



def load_contributions(
        file_path:str="../analytical-backend/data/08_reporting/dash_data_model.xlsx",
        ):

    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        assert "Local Explanation" in wb.sheetnames
        df = pd.read_excel(file_path, sheet_name="Local Explanation")

        df['Release Date'] = pd.to_datetime(df['Release Date'], utc=True).dt.strftime('%b %d')
        df["Data Series"] = df["Data Series"] + " (" + df["Series ID"] + ")"

        # Prepare the ordered dictionary

        df['Change'] = ['Up' if x > 0 else 'Down' if x < 0 else '' for x in df['Impact']]  # Adding 'Up' or 'Down' based on 'Impact'
        df["Impact"] = df["Impact"].map(lambda x: f"{x:.4f}")
        
        return df[['Release Date', "Data Series", "Impact", "Change"]].to_dict('records'), df["Series ID"].tolist()
    return None



# def load_series(series_id=None, diff=True):
#     def load_yaml(filepath):
#         with open(filepath, 'r') as file:
#             return yaml.load(file, Loader=yaml.FullLoader)
        
#     data_catalog = load_yaml(CATALOG_PATH)
#     parameters = load_yaml(PARAMETERS_PATH)

#     ref_datetime = pd.to_datetime(parameters['options']['ref_date'])
#     ref_date_col = parameters['options']['ref_date_col']
#     y_code = parameters['options']['y_code']
        
#     df = pd.read_excel(data_catalog['harmonized_data']['filepath'])
#     df[ref_date_col] = pd.to_datetime(df[ref_date_col])
#     df = df.sort_values(ref_date_col)

#     _, series_ids = load_contributions()
#     colnames = [ref_date_col, y_code] + series_ids

#     if diff:
#         for c in colnames[1:]:
#             df[c] = df[c].pct_change()
#     # df.loc[df[parameters['options']['ref_date_col']] == ref_datetime, y_code] = None

#     series = df.loc[df[ref_date_col].between(ref_datetime-relativedelta.relativedelta(months=7), ref_datetime-relativedelta.relativedelta(months=1)), colnames]
#     series["dt"] = pd.to_datetime(series[ref_date_col]).dt.strftime('%b')

#     if series_id is None:
#         return {
#             "labels": [month for month in series["dt"]],
#             "series": series[y_code].tolist()
#         }

#     return {
#         "labels": [month for month in series["dt"]],
#         "series": [series[c].tolist() for c in [y_code, series_id]]
#     }
def load_series(
        file_path:str="../analytical-backend/data/08_reporting/dash_data_model.xlsx",
        series_id=None,
        diff=True,
        ):
    def load_yaml(filepath):
        with open(filepath, 'r') as file:
            return yaml.load(file, Loader=yaml.FullLoader)
        
    # data_catalog = load_yaml(CATALOG_PATH)
    parameters = load_yaml(PARAMETERS_PATH)

    ref_datetime = pd.to_datetime(parameters['options']['ref_date'])
    ref_date_col = parameters['options']['ref_date_col']
    y_code = parameters['options']['y_code']
        
    # df = pd.read_excel(data_catalog['harmonized_data']['filepath'])
    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        assert "Global Explanation" in wb.sheetnames
        df = pd.read_excel(file_path, sheet_name="Global Explanation")

    df = df.pivot(index=ref_date_col, columns="Variable Code", values="Variable Value").reset_index()

    df[ref_date_col] = pd.to_datetime(df[ref_date_col])
    df = df.sort_values(ref_date_col)

    _, series_ids = load_contributions()
    colnames = [ref_date_col, y_code] + series_ids

    if diff:
        for c in colnames[1:]:
            df[c] = df[c].pct_change(fill_method=None)
    # df.loc[df[parameters['options']['ref_date_col']] == ref_datetime, y_code] = None

    series = df.loc[df[ref_date_col].between(ref_datetime-relativedelta.relativedelta(months=7), ref_datetime-relativedelta.relativedelta(months=1)), colnames]
    series["dt"] = pd.to_datetime(series[ref_date_col]).dt.strftime('%b')

    if series_id is None:
        return {
            "labels": [month for month in series["dt"]],
            "series": series[y_code].tolist()
        }

    return {
        "labels": [month for month in series["dt"]],
        "series": [series[c].tolist() for c in [y_code, series_id]]
    }

def load_evaluation(
        file_path:str="../analytical-backend/data/08_reporting/dash_data_model.xlsx",
        # type: str,
        ):
    items = {}
    if os.path.exists(file_path):
        wb = load_workbook(file_path, read_only=True)
        assert "Model Assessment" in wb.sheetnames
        df = pd.read_excel(file_path, sheet_name="Model Assessment")

        for k, v in zip(df["Measure"], df["Value"]):
            items[k] = v
    return items


# Global variables to store ...
# pipeline_status = "Not started"
# last_run_timestamp = "N/A"
TARGET_FOLDER = "../analytical-backend/data/_test"
PARAMETERS_PATH = "../analytical-backend/conf/base/parameters.yml"
CATALOG_PATH= "../analytical-backend/conf/base/catalog.yml"
parameters, data_catalog = __load_yaml(PARAMETERS_PATH), __load_yaml(CATALOG_PATH)

